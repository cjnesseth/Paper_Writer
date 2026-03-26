"""
01_parse_planning_params.py
Parse PJM BRA Planning Period Parameter files into a clean panel dataset.

Input:  Data/raw/planning_parameters/*.xlsx
Output: Data/cleaned/vrr_params.csv

VRR design split:
  Old (3-point, floor=$0):   2021/22 – 2025/26
  New (4-point, floor~$177): 2026/27 onward
"""

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

RAW_DIR     = Path("Data/raw/planning_parameters")
CLEANED_DIR = Path("Data/cleaned")
CLEANED_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# File manifest
# ---------------------------------------------------------------------------

FILE_MAP = [
    ("2021/22", "2021-2022-bra-planning-period-parameters.xlsx",                              "2021-2022 Parameters", "old"),
    ("2022/23", "2022-2023-planning-period-parameters-for-base-residual-auction.xlsx",        "Planning Parameters",  "old"),
    ("2023/24", "2023-2024-planning-period-parameters-for-base-residual-auction.xlsx",        "Planning Parameters",  "old"),
    ("2024/25", "2024-2025-rpm-bra-planning-parameters.xlsx",                                 "Planning Parameters",  "old"),
    ("2025/26", "2025-2026-planning-period-parameters-for-base-residual-auction.xlsx",        "Planning Parameters",  "old"),
    ("2026/27", "2026-2027-planning-period-parameters-for-base-residual-auction.xlsx",        "Planning Parameters",  "new"),
    ("2027/28", "2027-2028-planning-period-parameters-for-base-residual-auction.xlsx",        "Planning Parameters",  "new"),
    ("2028/29", "2028-2029-planning-period-parameters-for-base-residual-auction.xlsx",        "Planning Parameters",  "new"),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_float(val):
    """Convert a cell value to float, return None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def find_lda_header_row(ws):
    """Return 1-indexed row number of the LDA header row (col B = 'RTO', col C = 'MAAC')."""
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if isinstance(b, str) and b.strip() == "RTO" and \
           isinstance(c, str) and c.strip() == "MAAC":
            return r
    return None


def find_row_by_label(ws, pattern, start_row=1):
    """
    Return 1-indexed row of the first cell in column A matching `pattern`
    (case-insensitive regex), searching from start_row onward.
    """
    rx = re.compile(pattern, re.IGNORECASE)
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and rx.search(str(v)):
            return r
    return None


# ---------------------------------------------------------------------------
# Parse one file
# ---------------------------------------------------------------------------

def parse_file(delivery_year, filename, sheet_name, vrr_design):
    path = RAW_DIR / filename
    if not path.exists():
        print(f"  MISSING: {path}", file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Some files use the first sheet if the named sheet isn't found
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
        print(f"  WARNING: sheet '{sheet_name}' not found in {filename}; "
              f"using '{ws.title}'", file=sys.stderr)

    header_row = find_lda_header_row(ws)
    if header_row is None:
        print(f"  ERROR: LDA header row not found in {filename}", file=sys.stderr)
        wb.close()
        return []

    # Read LDA names and their column indices from the header row
    lda_cols = {}  # lda_name -> col_index (1-based)
    for col in range(2, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if v is not None and str(v).strip():
            lda_cols[str(v).strip()] = col

    # Locate key rows by label pattern (search below the header)
    def row(pattern):
        return find_row_by_label(ws, pattern, start_row=header_row + 1)

    rows = {
        "cetl":          row(r"CETL"),
        "rel_req":       row(r"Reliability Requirement.*FRR"),
        "net_cone":      row(r"Net CONE"),
        "pt_a_price":    row(r"Point\s*\(?a\)?.*Price|Price.*Point\s*\(?a\)?"),
        "pt_b_price":    row(r"Point\s*\(?b\)?.*Price|Price.*Point\s*\(?b\)?"),
        "pt_c_price":    row(r"Point\s*\(?c\)?.*Price|Price.*Point\s*\(?c\)?"),
        "pt_d_price":    row(r"Point\s*\(?d\)?.*Price|Price.*Point\s*\(?d\)?"),
        "pt_a_mw":       row(r"Point\s*\(?a\)?.*Level|Level.*Point\s*\(?a\)?"),
        "pt_b_mw":       row(r"Point\s*\(?b\)?.*Level|Level.*Point\s*\(?b\)?"),
        "pt_c_mw":       row(r"Point\s*\(?c\)?.*Level|Level.*Point\s*\(?c\)?"),
        "pt_d_mw":       row(r"Point\s*\(?d\)?.*Level|Level.*Point\s*\(?d\)?"),
    }

    def cell(row_key, col):
        r = rows[row_key]
        if r is None or col is None:
            return None
        return to_float(ws.cell(r, col).value)

    records = []
    for lda, col in lda_cols.items():
        records.append({
            "delivery_year":    delivery_year,
            "lda":              lda,
            "vrr_design":       vrr_design,
            "cetl_mw":          cell("cetl",       col),
            "reliability_req_mw": cell("rel_req",  col),
            "net_cone":         cell("net_cone",   col),
            "vrr_pt_a_price":   cell("pt_a_price", col),
            "vrr_pt_b_price":   cell("pt_b_price", col),
            "vrr_pt_c_price":   cell("pt_c_price", col),
            "vrr_pt_d_price":   cell("pt_d_price", col),
            "vrr_pt_a_mw":      cell("pt_a_mw",    col),
            "vrr_pt_b_mw":      cell("pt_b_mw",    col),
            "vrr_pt_c_mw":      cell("pt_c_mw",    col),
            "vrr_pt_d_mw":      cell("pt_d_mw",    col),
        })

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

all_records = []
for entry in FILE_MAP:
    delivery_year, filename, sheet_name, vrr_design = entry
    print(f"Parsing {delivery_year} …")
    records = parse_file(delivery_year, filename, sheet_name, vrr_design)
    print(f"  → {len(records)} LDA rows")
    all_records.extend(records)

df = pd.DataFrame(all_records)

# Spot-check
check = df.loc[(df["delivery_year"] == "2023/24") & (df["lda"] == "RTO"), "net_cone"]
if not check.empty and abs(check.iloc[0] - 274.96) < 0.1:
    print(f"\nSpot-check PASSED: 2023/24 RTO Net CONE = {check.iloc[0]}")
else:
    val = check.iloc[0] if not check.empty else "not found"
    print(f"\nSpot-check FAILED: 2023/24 RTO Net CONE = {val} (expected ~274.96)",
          file=sys.stderr)

out = CLEANED_DIR / "vrr_params.csv"
df.to_csv(out, index=False)
print(f"\nWrote {len(df)} rows to {out}")
