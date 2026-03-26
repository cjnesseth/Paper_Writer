"""
02_parse_bra_results.py
Parse PJM Base Residual Auction results files into a clean panel dataset.

Input:  Data/raw/bra_results/*.xlsx
Output: Data/cleaned/bra_clearing.csv

Sheet structures:
  2021/22  : unique layout — row 6 header; cols: LDA | MW | price
  2022/23+ : price rows start at row 5; "Participant Sell Offers Cleared"
             separator row divides prices from MW rows
"""

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

RAW_DIR     = Path("Data/raw/bra_results")
CLEANED_DIR = Path("Data/cleaned")
CLEANED_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# File manifest
# ---------------------------------------------------------------------------

FILE_MAP = [
    ("2021/22", "2021-2022-base-residual-auction-results.xlsx",  "2021"),
    ("2022/23", "2022-2023-base-residual-auction-results.xlsx",  "standard"),
    ("2023/24", "2023-2024-base-residual-auction-results.xlsx",  "standard"),
    ("2024/25", "2024-2025-base-residual-auction-results.xlsx",  "standard"),
    ("2025/26", "2025-2026-base-residual-auction-results.xlsx",  "standard"),
    ("2026/27", "2026-2027-bra-results.xlsx",                    "standard"),
    ("2027/28", "2027-2028-bra-results.xlsx",                    "standard"),
    # 2028/29 BRA results not yet available
]

# ---------------------------------------------------------------------------
# LDA name normalisation
# ---------------------------------------------------------------------------

LDA_RENAMES = {
    "PSNORTH":        "PS NORTH",
    "DPLSOUTH":       "DPL SOUTH",
    "ATSI-CLEVELAND": "ATSI-Cleveland",
}
# BRA results use "ATSI-CLEVELAND" (all-caps); planning params use "ATSI-Cleveland"

_NON_LDA = {"Total", "LDA", "Zone", ""}

def normalize_lda(raw):
    if raw is None:
        return None
    s = str(raw).strip().rstrip("*").strip()
    if s in _NON_LDA:
        return None
    return LDA_RENAMES.get(s, s)


def to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None

# ---------------------------------------------------------------------------
# Parse 2021/22 (unique layout)
# ---------------------------------------------------------------------------

def parse_2021_22(path):
    """
    Row 6 = header (LDA | MW cleared | price).
    Data rows start at row 7.
    Stop before the footnote row (col 1 starts with '*') to avoid
    reading the zonal-data section that follows.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    records = []
    for r in range(7, ws.max_row + 1):
        raw_lda = ws.cell(r, 1).value
        # Stop at footnote / section-header rows
        if raw_lda is None or str(raw_lda).strip().startswith("*"):
            break
        lda   = normalize_lda(raw_lda)
        mw    = to_float(ws.cell(r, 2).value)
        price = to_float(ws.cell(r, 3).value)
        if lda and price is not None:
            records.append({"lda": lda, "clearing_price": price, "mw_cleared": mw})

    wb.close()
    return records

# ---------------------------------------------------------------------------
# Parse standard layout (2022/23 – 2027/28)
# ---------------------------------------------------------------------------

def parse_standard(path):
    """
    Sheet "Summary".
    Rows 5+ = price rows (col 1 = LDA, col 2 = price) until separator row.
    Separator row col 1 contains "Participant Sell Offers Cleared".
    Rows after separator = MW rows (col 1 = LDA, col 2 = MW cleared).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    sheet_name = "Summary"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
        print(f"  WARNING: 'Summary' sheet not found; using '{ws.title}'",
              file=sys.stderr)

    prices = {}
    mw     = {}
    in_mw_section = False

    for r in range(5, ws.max_row + 1):
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value

        # Section separator: "Participant Buy Bids/Sell Offers Cleared" in col 1
        if col1 is not None and "Participant" in str(col1) and "Cleared" in str(col1):
            in_mw_section = True
            continue

        lda = normalize_lda(col1)
        if not lda:
            continue

        val = to_float(col2)
        if val is None:
            continue

        if in_mw_section:
            mw[lda] = val
        else:
            prices[lda] = val

    wb.close()

    # Merge prices and MW by LDA
    all_ldas = set(prices) | set(mw)
    return [
        {
            "lda":           lda,
            "clearing_price": prices.get(lda),
            "mw_cleared":    mw.get(lda),
        }
        for lda in sorted(all_ldas)
        if prices.get(lda) is not None or mw.get(lda) is not None
    ]

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

all_records = []
for delivery_year, filename, layout in FILE_MAP:
    path = RAW_DIR / filename
    if not path.exists():
        print(f"  MISSING: {path}", file=sys.stderr)
        continue

    print(f"Parsing {delivery_year} …")
    try:
        records = parse_2021_22(path) if layout == "2021" else parse_standard(path)
    except Exception as e:
        print(f"  ERROR in {filename}: {e}", file=sys.stderr)
        continue

    for rec in records:
        rec["delivery_year"] = delivery_year
    all_records.extend(records)
    print(f"  → {len(records)} LDA rows")

df = pd.DataFrame(all_records)[["delivery_year", "lda", "clearing_price", "mw_cleared"]]

# Spot-check
check = df.loc[(df["delivery_year"] == "2025/26") & (df["lda"] == "RTO"), "clearing_price"]
if not check.empty and abs(check.iloc[0] - 269.92) < 0.1:
    print(f"\nSpot-check PASSED: 2025/26 RTO clearing price = {check.iloc[0]}")
else:
    val = check.iloc[0] if not check.empty else "not found"
    print(f"\nSpot-check FAILED: 2025/26 RTO clearing price = {val} (expected 269.92)",
          file=sys.stderr)

out = CLEANED_DIR / "bra_clearing.csv"
df.to_csv(out, index=False)
print(f"\nWrote {len(df)} rows to {out}")
